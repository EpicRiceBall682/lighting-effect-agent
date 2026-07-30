"""Export successful batch results into the organizer's XLSX template."""

from __future__ import annotations

import argparse
from copy import copy
from io import BytesIO
import json
from pathlib import Path
import sys
from typing import Any


PROJECT_ROOT = Path(__file__).resolve().parents[3]
DEFAULT_TEMPLATE = PROJECT_ROOT / "reference_data" / "测试集提交格式.xlsx"
EXPECTED_HEADERS = ("English Prompt", "中文 Prompt", "Image", "Scene")


def load_batch_results(path: Path) -> list[dict[str, Any]]:
    by_case_id: dict[str, dict[str, Any]] = {}
    for line_number, line in enumerate(Path(path).read_text(encoding="utf-8").splitlines(), 1):
        if not line.strip():
            continue
        try:
            record = json.loads(line)
        except json.JSONDecodeError as exc:
            raise ValueError(f"{path}:{line_number} is not valid JSONL") from exc
        case_id = str(record.get("case_id", ""))
        if not case_id:
            raise ValueError(f"{path}:{line_number} has no case_id")
        # Batch retries append a newer record instead of rewriting history.
        by_case_id[case_id] = record
    records = list(by_case_id.values())
    records.sort(key=lambda item: int(item.get("case_index", 0)))
    return records


def export_submission(
    results_path: Path,
    template_path: Path,
    output_path: Path,
    *,
    image_field: str = "raw_image_path",
    allow_incomplete: bool = False,
) -> dict[str, Any]:
    """Fill all four submission columns and embed one image per successful case."""

    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as SpreadsheetImage
    except ImportError as exc:
        raise RuntimeError(
            "Excel export requires openpyxl>=3.1; install module-6 requirements"
        ) from exc

    records = load_batch_results(results_path)
    failures = [record for record in records if record.get("status") != "success"]
    if failures and not allow_incomplete:
        case_ids = ", ".join(str(item.get("case_id", "?")) for item in failures)
        raise ValueError(f"batch contains failed cases; refusing incomplete export: {case_ids}")
    successful = [record for record in records if record.get("status") == "success"]
    if not successful:
        raise ValueError("batch contains no successful records")

    workbook = load_workbook(template_path)
    worksheet = workbook["prompts"] if "prompts" in workbook.sheetnames else workbook.active
    headers = tuple(worksheet.cell(1, column).value for column in range(1, 5))
    if headers != EXPECTED_HEADERS:
        raise ValueError(
            f"unexpected submission headers: {headers!r}; expected {EXPECTED_HEADERS!r}"
        )

    if worksheet.max_row > 1:
        worksheet.delete_rows(2, worksheet.max_row - 1)
    # Remove the example drawings shipped with the organizer template.
    worksheet._images.clear()
    worksheet.freeze_panes = "A2"
    image_buffers: list[BytesIO] = []

    for row_number, record in enumerate(successful, 2):
        image_path = Path(str(record.get(image_field, ""))).expanduser()
        if not image_path.is_file():
            raise FileNotFoundError(
                f"{image_field} does not exist for {record.get('case_id')}: {image_path}"
            )
        worksheet.cell(row_number, 1, str(record.get("english_prompt", "")))
        worksheet.cell(
            row_number,
            2,
            str(record.get("chinese_prompt") or record.get("scene", "")),
        )
        worksheet.cell(row_number, 4, str(record.get("scene", "")))
        worksheet.row_dimensions[row_number].height = 90
        for column in (1, 2, 4):
            cell = worksheet.cell(row_number, column)
            alignment = copy(cell.alignment)
            alignment.wrap_text = True
            alignment.vertical = "center"
            cell.alignment = alignment

        image_buffer = BytesIO(image_path.read_bytes())
        image_buffers.append(image_buffer)
        picture = SpreadsheetImage(image_buffer)
        max_width, max_height = 240, 112
        scale = min(max_width / picture.width, max_height / picture.height)
        picture.width = round(picture.width * scale)
        picture.height = round(picture.height * scale)
        picture.anchor = f"C{row_number}"
        worksheet.add_image(picture)

    output_path = Path(output_path).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    for image_buffer in image_buffers:
        image_buffer.close()
    return {
        "output_path": str(output_path),
        "exported_count": len(successful),
        "skipped_failed_count": len(failures),
        "image_field": image_field,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Export batch results to the organizer XLSX.")
    parser.add_argument("--results", type=Path, required=True)
    parser.add_argument("--template", type=Path, default=DEFAULT_TEMPLATE)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument(
        "--image-field",
        choices=("raw_image_path", "sdl_preview_path", "sdl_control_path"),
        default="raw_image_path",
    )
    parser.add_argument("--allow-incomplete", action="store_true")
    args = parser.parse_args(argv)
    try:
        summary = export_submission(
            args.results,
            args.template,
            args.output,
            image_field=args.image_field,
            allow_incomplete=args.allow_incomplete,
        )
    except (OSError, RuntimeError, ValueError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        return 1
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
