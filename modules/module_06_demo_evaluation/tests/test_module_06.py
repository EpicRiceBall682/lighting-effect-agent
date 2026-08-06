from __future__ import annotations

from dataclasses import dataclass
import importlib.util
import json
from pathlib import Path
from types import SimpleNamespace
import tempfile
import unittest
import zipfile

import numpy as np
from PIL import Image

from modules.module_01_prompt_agent.src.schemas import LightingEffectAttributes
from modules.module_03_image_generation.src.generator import GenerationResult
from modules.module_04_gamut_mapping.tests.sample_palette import (
    write_sample_sdl_palette,
)
from modules.module_06_demo_evaluation.src.app import (
    _format_metrics,
    build_demo,
    build_parser,
)
from modules.module_06_demo_evaluation.src.evaluator import (
    BatchEvaluationConfig,
    load_test_scenes,
    run_batch_evaluation,
)
from modules.module_06_demo_evaluation.src.exporter import (
    export_submission,
    load_batch_results,
)
from modules.module_06_demo_evaluation.src.pipeline import (
    LightingDemoPipeline,
    image_mean_absolute_difference,
    scene_aware_seed,
    scene_palette_notice,
)


PROJECT_ROOT = Path(__file__).resolve().parents[3]


class WindowsLauncherTests(unittest.TestCase):
    def test_launcher_prefers_cuda_and_passes_selected_device(self):
        launcher = (PROJECT_ROOT / "start_demo_windows.bat").read_text(
            encoding="utf-8"
        )

        self.assertIn("torch.cuda.is_available()", launcher)
        self.assertIn("torch.cuda.get_device_name(0)", launcher)
        self.assertIn('--device "%DEVICE%"', launcher)
        self.assertIn("https://pytorch.org/get-started/locally/", launcher)


class FakePromptAgent:
    def generate(self, scene: str, **kwargs):
        return LightingEffectAttributes(
            density="middle",
            m_intensity=70,
            k_intensity=55,
            a_intensity=65,
            effect=(
                "Wide panoramic abstract light texture with warm amber at the center, "
                "pale orange spreading toward both sides, a broad horizontal gradient, "
                "gentle misty glow, diffused bloom, evenly bright illumination, and a "
                "welcoming coffee atmosphere."
            ),
        )


class FakeGenerator:
    construction_count = 0
    generation_count = 0
    generated_seeds: list[int] = []

    def __init__(self, **kwargs):
        type(self).construction_count += 1

    def generate(self, prompt, *, output_dir, config, source_attributes):
        type(self).generation_count += 1
        type(self).generated_seeds.append(config.seed)
        output_dir.mkdir(parents=True, exist_ok=True)
        x = np.linspace(0, 1, config.width, dtype=np.float32)
        array = np.empty((config.height, config.width, 3), dtype=np.uint8)
        array[:, :, 0] = np.rint(150 + 105 * x)
        array[:, :, 1] = np.rint(80 + 130 * x)
        array[:, :, 2] = np.rint(190 - 120 * x)
        image_path = output_dir / f"raw_light_effect_seed_{config.seed}.png"
        manifest_path = image_path.with_suffix(".json")
        Image.fromarray(array, mode="RGB").save(image_path)
        manifest_path.write_text(
            json.dumps({"prompt": prompt, "seed": config.seed}) + "\n",
            encoding="utf-8",
        )
        return GenerationResult(
            image_path=image_path,
            manifest_path=manifest_path,
            seed=config.seed,
            width=config.width,
            height=config.height,
        )

    def generate_concept(self, prompt, *, output_dir, seed, steps):
        output_dir.mkdir(parents=True, exist_ok=True)
        array = np.zeros((288, 512, 3), dtype=np.uint8)
        array[:, :128] = (245, 150, 190)
        array[:, 128:256] = (245, 220, 100)
        array[:, 256:384] = (70, 205, 105)
        array[:, 384:] = (80, 185, 225)
        image_path = output_dir / f"concept_image_seed_{seed}.png"
        manifest_path = output_dir / "concept_image.json"
        Image.fromarray(array, mode="RGB").save(image_path)
        manifest_path.write_text(
            json.dumps({"prompt": prompt, "steps": steps}) + "\n",
            encoding="utf-8",
        )
        return SimpleNamespace(
            image_path=image_path,
            manifest_path=manifest_path,
            seed=seed,
            steps=steps,
            inference_seconds=0.01,
        )


class SamplePaletteTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls._palette_directory = tempfile.TemporaryDirectory()
        cls.sample_sdl_path = Path(cls._palette_directory.name) / "sample_sdl.txt"
        write_sample_sdl_palette(cls.sample_sdl_path)

    @classmethod
    def tearDownClass(cls):
        cls._palette_directory.cleanup()


class PipelineTests(SamplePaletteTestCase):
    def setUp(self):
        FakeGenerator.construction_count = 0
        FakeGenerator.generation_count = 0
        FakeGenerator.generated_seeds = []

    def test_scene_aware_seed_is_stable_but_changes_between_scenes(self):
        first = scene_aware_seed("蓝天白云大草原", 20260724)
        repeated = scene_aware_seed("  蓝天白云大草原  ", 20260724)
        second = scene_aware_seed("沙滩和大海", 20260724)
        self.assertEqual(first, repeated)
        self.assertNotEqual(first, second)

    def test_image_difference_is_zero_for_identical_images(self):
        first = Image.new("RGB", (64, 32), (120, 150, 180))
        second = Image.new("RGB", (128, 64), (120, 150, 180))
        self.assertEqual(image_mean_absolute_difference(first, second), 0.0)

    def test_pipeline_writes_complete_artifact_bundle_and_reuses_generator(self):
        with tempfile.TemporaryDirectory() as directory:
            pipeline = LightingDemoPipeline(
                sdl_path=self.sample_sdl_path,
                output_root=Path(directory),
                prompt_agent=FakePromptAgent(),
                generator_factory=FakeGenerator,
            )
            result = pipeline.run(
                "温暖惬意的咖啡厅",
                1220,
                370,
                80,
                seed=42,
                steps=10,
            )
            second = pipeline.run(
                "清晨柔和的酒店大堂",
                1220,
                370,
                seed=43,
                steps=10,
            )

            self.assertEqual(FakeGenerator.construction_count, 1)
            self.assertEqual(result.width, 1024)
            self.assertEqual(result.height % 8, 0)
            self.assertEqual(result.quality["strict_invalid_pixel_count"], 0)
            for path in (
                result.raw_image_path,
                result.themed_image_path,
                result.sdl_preview_path,
                result.sdl_control_path,
                result.out_of_gamut_mask_path,
                result.prompt_json_path,
                result.generation_manifest_path,
                result.report_path,
                result.archive_path,
                second.archive_path,
            ):
                self.assertTrue(path.is_file(), path)
            with zipfile.ZipFile(result.archive_path) as archive:
                names = set(archive.namelist())
            self.assertIn("module_01_prompt.json", names)
            self.assertIn("module_05_pattern.json", names)
            self.assertIn(result.themed_image_path.name, names)
            self.assertIn("pipeline_report.json", names)
            self.assertIn(result.raw_image_path.name, names)

    def test_changed_scene_retries_when_result_is_perceptually_duplicate(self):
        with tempfile.TemporaryDirectory() as directory:
            pipeline = LightingDemoPipeline(
                sdl_path=self.sample_sdl_path,
                output_root=Path(directory),
                prompt_agent=FakePromptAgent(),
                generator_factory=FakeGenerator,
            )
            pipeline.run(
                "蓝天白云大草原",
                1220,
                370,
                seed=20260724,
                steps=10,
            )
            second = pipeline.run(
                "沙滩和大海",
                1220,
                370,
                seed=20260724,
                steps=10,
            )
            report = json.loads(second.report_path.read_text(encoding="utf-8"))
            guard = report["module_03"]["similarity_guard"]

            self.assertEqual(second.seed_mode, "scene_derived")
            self.assertEqual(second.similarity_retry_count, 1)
            self.assertEqual(FakeGenerator.generation_count, 3)
            self.assertEqual(guard["retry_count"], 1)
            self.assertEqual(guard["initial_mean_absolute_difference"], 0.0)
            self.assertNotEqual(second.effective_seed, second.requested_seed)

    def test_module_five_is_between_raw_generation_and_module_four(self):
        with tempfile.TemporaryDirectory() as directory:
            pipeline = LightingDemoPipeline(
                sdl_path=self.sample_sdl_path,
                output_root=Path(directory),
                prompt_agent=FakePromptAgent(),
                generator_factory=FakeGenerator,
            )
            result = pipeline.run(
                "温暖惬意的咖啡厅",
                1220,
                370,
                seed=42,
                steps=10,
            )
            report = json.loads(result.report_path.read_text(encoding="utf-8"))
            self.assertEqual(
                report["module_05"]["input_raw_sha256"],
                report["module_03"]["raw_image_sha256"],
            )
            self.assertEqual(
                report["module_04"]["input_raw_sha256"],
                report["module_05"]["themed_image_sha256"],
            )
            self.assertEqual(
                report["module_04"]["input_source"],
                "module_05_themed_image",
            )

    def test_module_five_can_be_safely_disabled(self):
        with tempfile.TemporaryDirectory() as directory:
            pipeline = LightingDemoPipeline(
                sdl_path=self.sample_sdl_path,
                output_root=Path(directory),
                prompt_agent=FakePromptAgent(),
                generator_factory=FakeGenerator,
            )
            result = pipeline.run(
                "温暖惬意的咖啡厅",
                1220,
                370,
                seed=42,
                steps=10,
                pattern_enabled=False,
            )
            with (
                Image.open(result.raw_image_path) as raw,
                Image.open(result.themed_image_path) as themed,
            ):
                self.assertEqual(raw.convert("RGB").tobytes(), themed.convert("RGB").tobytes())
            self.assertEqual(result.pattern_report["quality_status"], "bypassed")

    def test_sdl_quality_failure_retries_generation_once_and_records_traceability(self):
        from dataclasses import replace

        from modules.module_04_gamut_mapping.src.mapper import GamutMapper

        class FailOnceMapper:
            calls = 0

            def __init__(self, palette):
                self.mapper = GamutMapper(palette)

            def map_image(self, image, *, method):
                type(self).calls += 1
                result = self.mapper.map_image(image, method=method)
                if type(self).calls == 1:
                    return replace(
                        result,
                        quality_failures=(
                            "before_xy_out_of_gamut_fraction=0.500000 "
                            "violates its maximum 0.300000",
                        ),
                    )
                return result

        with tempfile.TemporaryDirectory() as directory:
            FailOnceMapper.calls = 0
            pipeline = LightingDemoPipeline(
                sdl_path=self.sample_sdl_path,
                output_root=Path(directory),
                prompt_agent=FakePromptAgent(),
                generator_factory=FakeGenerator,
                mapper_factory=FailOnceMapper,
            )
            result = pipeline.run(
                "运动空间的抽象能量光",
                1220,
                370,
                seed=45,
                steps=10,
            )
            report = json.loads(result.report_path.read_text(encoding="utf-8"))

            self.assertEqual(result.sdl_retry_count, 1)
            self.assertEqual(FakeGenerator.generation_count, 2)
            self.assertEqual(report["module_04"]["retry_count"], 1)
            self.assertEqual(len(report["module_04"]["attempts"]), 2)
            self.assertTrue(report["module_04"]["attempts"][0]["failures"])
            self.assertFalse(report["module_04"]["attempts"][1]["failures"])
            self.assertEqual(
                report["traceability"]["lora_sha256"],
                "32140f4d8750e8b6b43f6440e7e28fa7ab5bb7840de68f36fb4733c81ba2ddd0",
            )
            self.assertIn(
                "dataset_zip_sha256",
                report["traceability"]["weight_provenance"],
            )
            self.assertIn("module_05", report)

    def test_fixed_seed_mode_uses_exact_user_seed(self):
        with tempfile.TemporaryDirectory() as directory:
            pipeline = LightingDemoPipeline(
                sdl_path=self.sample_sdl_path,
                output_root=Path(directory),
                prompt_agent=FakePromptAgent(),
                generator_factory=FakeGenerator,
            )
            result = pipeline.run(
                "温暖惬意的咖啡厅",
                1220,
                370,
                seed=42,
                steps=10,
                fixed_seed=True,
            )
            self.assertEqual(result.effective_seed, 42)
            self.assertEqual(result.seed_mode, "fixed")

    def test_fixed_seed_is_not_changed_by_similarity_retry(self):
        with tempfile.TemporaryDirectory() as directory:
            pipeline = LightingDemoPipeline(
                sdl_path=self.sample_sdl_path,
                output_root=Path(directory),
                prompt_agent=FakePromptAgent(),
                generator_factory=FakeGenerator,
            )
            pipeline.run(
                "第一个足够长的固定种子场景",
                1220,
                370,
                seed=42,
                steps=10,
                fixed_seed=True,
            )
            second = pipeline.run(
                "第二个足够长的固定种子场景",
                1220,
                370,
                seed=42,
                steps=10,
                fixed_seed=True,
            )
            self.assertEqual(second.effective_seed, 42)
            self.assertEqual(second.similarity_retry_count, 0)

    def test_failed_run_removes_incomplete_run_directory(self):
        with tempfile.TemporaryDirectory() as directory:
            pipeline = LightingDemoPipeline(
                sdl_path=self.sample_sdl_path,
                output_root=Path(directory),
                prompt_agent=FakePromptAgent(),
                generator_factory=FakeGenerator,
            )
            with self.assertRaisesRegex(ValueError, "结构化参数"):
                pipeline.run(
                    "足够长的场景描述",
                    1220,
                    370,
                    prompt_override="edited prompt",
                    attributes_override=None,
                )
            self.assertEqual(list(Path(directory).iterdir()), [])

    def test_input_validation_happens_before_creating_a_run_directory(self):
        with tempfile.TemporaryDirectory() as directory:
            pipeline = LightingDemoPipeline(
                sdl_path=self.sample_sdl_path,
                output_root=Path(directory),
                prompt_agent=FakePromptAgent(),
                generator_factory=FakeGenerator,
            )
            with self.assertRaisesRegex(ValueError, "灯具"):
                pipeline.run("足够长的场景描述", 0, 370)
            self.assertEqual(list(Path(directory).iterdir()), [])

    def test_nonfinite_input_is_rejected_before_creating_a_run_directory(self):
        with tempfile.TemporaryDirectory() as directory:
            pipeline = LightingDemoPipeline(
                sdl_path=self.sample_sdl_path,
                output_root=Path(directory),
                prompt_agent=FakePromptAgent(),
                generator_factory=FakeGenerator,
            )
            with self.assertRaisesRegex(ValueError, "灯具"):
                pipeline.run("足够长的场景描述", float("nan"), 370)
            self.assertEqual(list(Path(directory).iterdir()), [])

    def test_green_scene_requires_no_palette_translation(self):
        notice = scene_palette_notice("蓝天白云大草原")
        self.assertEqual(notice, "")

    def test_missing_sdl_table_uses_explicit_preview_mode(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            missing_sdl = root / "private" / "SDL2_0.txt"
            pipeline = LightingDemoPipeline(
                sdl_path=missing_sdl,
                output_root=root / "outputs",
                prompt_agent=FakePromptAgent(),
                generator_factory=FakeGenerator,
            )
            result = pipeline.run(
                "Windows 本地预览模式场景",
                1220,
                370,
                seed=42,
                steps=10,
            )
            report = json.loads(result.report_path.read_text(encoding="utf-8"))

            self.assertFalse(result.sdl_available)
            self.assertEqual(
                result.quality["status"],
                "skipped_missing_sdl_table",
            )
            self.assertTrue(result.sdl_preview_path.is_file())
            self.assertTrue(result.sdl_control_path.is_file())
            self.assertTrue(result.out_of_gamut_mask_path.is_file())
            self.assertEqual(
                report["module_04"]["quality_status"],
                "skipped_missing_sdl_table",
            )
            self.assertIsNone(report["traceability"]["sdl_table_sha256"])

    def test_missing_sdl_table_can_be_required(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            pipeline = LightingDemoPipeline(
                sdl_path=root / "missing.txt",
                output_root=root,
                prompt_agent=FakePromptAgent(),
                generator_factory=FakeGenerator,
                allow_missing_sdl=False,
            )
            with self.assertRaisesRegex(FileNotFoundError, "SDL color table"):
                pipeline.run(
                    "必须使用 SDL 的测试场景",
                    1220,
                    370,
                    seed=42,
                    steps=10,
                )
            self.assertEqual(list(root.iterdir()), [])

    def test_edited_prompt_rerun_skips_api_contract_and_is_recorded(self):
        with tempfile.TemporaryDirectory() as directory:
            pipeline = LightingDemoPipeline(
                sdl_path=self.sample_sdl_path,
                output_root=Path(directory),
                prompt_agent=FakePromptAgent(),
                generator_factory=FakeGenerator,
            )
            attributes = FakePromptAgent().generate("test").to_dict()
            edited = (
                "Wide panoramic abstract light texture with pale yellow across the lower "
                "area, light blue above, a broad vertical transition, cloud-like mist, "
                "soft diffused bloom, evenly bright illumination, and a spacious airy "
                "atmosphere throughout the panel."
            )
            result = pipeline.run(
                "蓝天白云大草原",
                1220,
                370,
                seed=44,
                steps=10,
                prompt_override=edited,
                attributes_override=attributes,
            )
            report = json.loads(result.report_path.read_text(encoding="utf-8"))
            self.assertEqual(result.prompt, edited)
            self.assertEqual(report["module_01_prompt_source"], "user_edited")
            self.assertEqual(result.palette_notice, "")

    def test_fast_mode_returns_concept_and_green_light_within_budget(self):
        with tempfile.TemporaryDirectory() as directory:
            pipeline = LightingDemoPipeline(
                sdl_path=Path(directory) / "missing_sdl.txt",
                output_root=Path(directory) / "outputs",
                generator_factory=FakeGenerator,
                fast_mode=True,
                time_budget_seconds=6.0,
            )
            result = pipeline.run(
                "清晨花海，粉色、黄色、绿色和蓝色，清新自然",
                1220,
                370,
                seed=42,
                steps=4,
            )
            self.assertTrue(result.concept_image_path.is_file())
            self.assertTrue(result.sdl_preview_path.is_file())
            self.assertNotIn("derived from the concept scene palette", result.prompt)
            self.assertTrue(result.deadline_met)
            self.assertLess(result.timings["total_seconds"], 6.0)
            report = json.loads(result.report_path.read_text(encoding="utf-8"))
            self.assertEqual(
                report["concept_image"]["path"],
                report["concept_image"]["source_path"],
            )
            concept_manifest = json.loads(
                result.concept_manifest_path.read_text(encoding="utf-8")
            )
            self.assertFalse(concept_manifest["harmonization"]["applied"])
            self.assertEqual(
                report["module_03"]["prompt_color_guidance"]["palette_source"],
                "module_01_independent_effect_prompt",
            )
            self.assertFalse(
                report["module_03"]["prompt_color_guidance"][
                    "correction_applied"
                ]
            )

    def test_fast_mode_uses_model_for_the_natural_concept_prompt(self):
        class AutoPaletteAgent:
            calls = 0

            def generate(self, _scene, **_kwargs):
                type(self).calls += 1
                return LightingEffectAttributes(
                    density="middle",
                    m_intensity=82,
                    k_intensity=90,
                    a_intensity=74,
                    effect=(
                        "Wide panoramic color field with electric purple on the left, vivid "
                        "magenta through the center, and dominant bright cyan across the "
                        "right, forming a smooth horizontal gradient with uniform vertical "
                        "color, futuristic neon illumination, visual coherence, and an "
                        "uninterrupted luminous surface throughout."
                    ),
                    concept_prompt=(
                        "Cinematic futuristic neon city street with reflective surfaces, "
                        "layered architecture, electric purple, vivid magenta, and bright "
                        "cyan lighting, realistic detail and depth."
                    ),
                )

        with tempfile.TemporaryDirectory() as directory:
            agent = AutoPaletteAgent()
            pipeline = LightingDemoPipeline(
                sdl_path=Path(directory) / "missing_sdl.txt",
                output_root=Path(directory) / "outputs",
                generator_factory=FakeGenerator,
                fast_mode=True,
                auto_palette_agent_factory=lambda: agent,
            )
            result = pipeline.run("赛博朋克", 1220, 370, seed=42, steps=4)
            report = json.loads(result.report_path.read_text(encoding="utf-8"))

            self.assertEqual(result.prompt_source, "deepseek_concept_prompt")
            self.assertEqual(agent.calls, 1)
            self.assertIn("electric purple", result.prompt)
            self.assertNotIn("derived from the concept scene palette", result.prompt)
            self.assertIn(
                "futuristic neon city street",
                result.attributes["concept_prompt"],
            )
            self.assertEqual(
                report["module_01_auto_palette"]["reason"],
                "model_generated_natural_concept_prompt",
            )

    def test_explicit_color_also_uses_model_palette_judgment(self):
        class ExplicitColorAgent:
            calls = 0

            def generate(self, _scene, **_kwargs):
                type(self).calls += 1
                return LightingEffectAttributes(
                    density="middle",
                    m_intensity=78,
                    k_intensity=84,
                    a_intensity=70,
                    effect=(
                        "Wide panoramic color field with dominant bright green across the "
                        "entire panel, forming a clean smooth horizontal gradient with "
                        "uniform vertical color, balanced illumination, natural clarity, "
                        "visual coherence, and an uninterrupted luminous surface throughout."
                    ),
                    concept_prompt=(
                        "Cinematic forest clearing with layered trees and bright green "
                        "natural light, realistic detail, depth, and an open composition."
                    ),
                )

        with tempfile.TemporaryDirectory() as directory:
            agent = ExplicitColorAgent()
            pipeline = LightingDemoPipeline(
                sdl_path=Path(directory) / "missing_sdl.txt",
                output_root=Path(directory) / "outputs",
                generator_factory=FakeGenerator,
                fast_mode=True,
                auto_palette_agent_factory=lambda: agent,
            )
            result = pipeline.run("绿色森林", 1220, 370, seed=42, steps=4)

            self.assertEqual(result.prompt_source, "deepseek_concept_prompt")
            self.assertEqual(agent.calls, 1)
            self.assertIn("dominant bright green", result.prompt)
            self.assertNotIn("derived from the concept scene palette", result.prompt)

    def test_failed_model_palette_uses_semantic_style_fallback(self):
        class FailingAgent:
            def generate(self, _scene, **_kwargs):
                raise RuntimeError("offline")

        with tempfile.TemporaryDirectory() as directory:
            pipeline = LightingDemoPipeline(
                sdl_path=Path(directory) / "missing_sdl.txt",
                output_root=Path(directory) / "outputs",
                generator_factory=FakeGenerator,
                fast_mode=True,
                auto_palette_agent_factory=FailingAgent,
            )
            result = pipeline.run("赛博朋克", 1220, 370, seed=42, steps=4)
            report = json.loads(result.report_path.read_text(encoding="utf-8"))

            self.assertEqual(result.prompt_source, "local_concept_prompt_fallback")
            self.assertIn("electric purple", result.prompt)
            self.assertNotIn("derived from the concept scene palette", result.prompt)
            self.assertIn("futuristic city", result.attributes["concept_prompt"])
            self.assertEqual(
                report["module_01_auto_palette"]["reason"],
                "unexpected_model_failure",
            )

    def test_severe_independent_color_mismatch_uses_concept_palette_fallback(self):
        class RedConceptGenerator(FakeGenerator):
            def generate_concept(self, prompt, *, output_dir, seed, steps):
                output_dir.mkdir(parents=True, exist_ok=True)
                image_path = output_dir / f"concept_image_seed_{seed}.png"
                manifest_path = output_dir / "concept_image.json"
                Image.new("RGB", (512, 288), (240, 45, 35)).save(image_path)
                manifest_path.write_text(
                    json.dumps({"prompt": prompt, "steps": steps}) + "\n",
                    encoding="utf-8",
                )
                return SimpleNamespace(
                    image_path=image_path,
                    manifest_path=manifest_path,
                    seed=seed,
                    steps=steps,
                    inference_seconds=0.01,
                )

        class BlueEffectAgent:
            def generate(self, _scene, **_kwargs):
                return LightingEffectAttributes(
                    density="middle",
                    m_intensity=75,
                    k_intensity=82,
                    a_intensity=68,
                    effect=(
                        "Wide panoramic color field with dominant bright blue across the "
                        "entire panel, forming a clean smooth horizontal gradient with "
                        "uniform vertical color, balanced illumination, natural clarity, "
                        "visual coherence, and an uninterrupted luminous surface throughout."
                    ),
                    concept_prompt=(
                        "A vivid red exhibition space with realistic materials, people, "
                        "architectural depth, and strongly saturated environmental light."
                    ),
                )

        with tempfile.TemporaryDirectory() as directory:
            pipeline = LightingDemoPipeline(
                sdl_path=Path(directory) / "missing_sdl.txt",
                output_root=Path(directory) / "outputs",
                generator_factory=RedConceptGenerator,
                fast_mode=True,
                auto_palette_agent_factory=BlueEffectAgent,
            )
            result = pipeline.run("强烈红色展览空间", 1220, 370, seed=42, steps=4)
            report = json.loads(result.report_path.read_text(encoding="utf-8"))
            guidance = report["module_03"]["prompt_color_guidance"]

            self.assertTrue(guidance["correction_applied"])
            self.assertTrue(guidance["color_comparison"]["correction_required"])
            self.assertEqual(
                guidance["palette_source"],
                "concept_image_fallback_large_color_mismatch",
            )
            self.assertIn("derived from the concept scene palette", result.prompt)


class AppTests(SamplePaletteTestCase):
    def test_metrics_expose_dynamic_brightness_floor(self):
        metrics = _format_metrics(
            {
                "mapping_available": False,
                "pixel_count": 100,
            },
            color_guidance={
                "brightness_floor": {
                    "applied": True,
                    "policy": {"mode": "energetic"},
                    "before": {"mean_luminance": 0.145},
                    "after": {
                        "mean_luminance": 0.32,
                        "below_0_20_fraction": 0.0,
                    },
                }
            },
        )

        self.assertEqual(metrics["Raw 亮度策略"], "活力/霓虹场景")
        self.assertEqual(metrics["Raw 自动增亮"], "已启用")
        self.assertEqual(metrics["Raw 平均亮度（最终）"], 0.32)

    def test_gradio_app_builds_without_loading_the_diffusion_model(self):
        with tempfile.TemporaryDirectory() as directory:
            pipeline = LightingDemoPipeline(
                sdl_path=self.sample_sdl_path,
                output_root=Path(directory),
                prompt_agent=FakePromptAgent(),
                generator_factory=FakeGenerator,
            )
            demo = build_demo(pipeline)
            self.assertGreater(len(demo.blocks), 10)
            self.assertEqual(FakeGenerator.construction_count, 0)

    def test_windows_friendly_sdl_cli_options(self):
        args = build_parser().parse_args(
            ["--sdl-path", r"C:\private\SDL2_0.txt", "--require-sdl"]
        )
        self.assertEqual(str(args.sdl_path), r"C:\private\SDL2_0.txt")
        self.assertTrue(args.require_sdl)
        self.assertEqual(args.auto_palette_timeout_seconds, 3.0)

    def test_metrics_do_not_claim_sdl_compliance_in_preview_mode(self):
        metrics = _format_metrics(
            {
                "mapping_available": False,
                "status": "skipped_missing_sdl_table",
                "pixel_count": 1024,
            }
        )
        self.assertIn("未运行", metrics["SDL 映射状态"])
        self.assertIn("无", metrics["硬件色域保证"])


class BatchEvaluationTests(unittest.TestCase):
    class FakeBatchPipeline:
        def __init__(self, root: Path, failing_scene: str | None = None):
            self.root = root
            self.failing_scene = failing_scene
            self.calls = 0
            self.forbidden_prompt_calls = []

        def run(self, scene, width_mm, height_mm, space_size_m2, **kwargs):
            self.calls += 1
            self.forbidden_prompt_calls.append(tuple(kwargs.get("forbidden_prompts", ())))
            if scene == self.failing_scene:
                raise RuntimeError("intentional test failure")
            run_dir = self.root / f"run-{self.calls}"
            run_dir.mkdir(parents=True)
            image_path = run_dir / "raw.png"
            Image.new("RGB", (96, 32), (240, 180, 100)).save(image_path)
            report_path = run_dir / "report.json"
            report_path.write_text("{}\n", encoding="utf-8")
            archive_path = run_dir / "result.zip"
            archive_path.write_bytes(b"zip")
            return SimpleNamespace(
                prompt=f"English lighting prompt for {scene}",
                attributes={"density": "middle"},
                raw_image_path=image_path,
                sdl_preview_path=image_path,
                sdl_control_path=image_path,
                report_path=report_path,
                archive_path=archive_path,
                effective_seed=kwargs["seed"],
                seed_mode="scene_derived",
                raw_quality={"mean_luminance": 0.7},
                quality={"strict_invalid_pixel_count": 0},
                similarity_retry_count=0,
                similarity_difference=None,
            )

    def test_test_set_loader_preserves_nonempty_scenes(self):
        with tempfile.TemporaryDirectory() as directory:
            test_set = Path(directory) / "test_scenes.txt"
            test_set.write_text("场景一\n\n 场景二 \n", encoding="utf-8")
            self.assertEqual(load_test_scenes(test_set), ["场景一", "场景二"])

    def test_batch_is_resumable_and_records_each_case(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            pipeline = self.FakeBatchPipeline(root / "runs")
            scenes = ["酒店大堂晨雾天光", "酒店客房黄昏助眠光"]
            first = run_batch_evaluation(
                pipeline,
                scenes,
                root / "batch",
                config=BatchEvaluationConfig(steps=8),
            )
            second = run_batch_evaluation(
                pipeline,
                scenes,
                root / "batch",
                config=BatchEvaluationConfig(steps=8),
            )
            records = load_batch_results(root / "batch" / "batch_results.jsonl")

            self.assertTrue(first["complete"])
            self.assertEqual(second["reused_successes"], 2)
            self.assertEqual(pipeline.calls, 2)
            self.assertEqual(len(records), 2)
            self.assertEqual(records[0]["chinese_prompt"], scenes[0])
            self.assertEqual(pipeline.forbidden_prompt_calls[0], ())
            self.assertEqual(
                pipeline.forbidden_prompt_calls[1],
                (f"English lighting prompt for {scenes[0]}",),
            )

    def test_batch_resume_invalidates_success_when_config_changes(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            pipeline = self.FakeBatchPipeline(root / "runs")
            scenes = ["同一个批量评估场景"]
            run_batch_evaluation(
                pipeline,
                scenes,
                root / "batch",
                config=BatchEvaluationConfig(seed=1, steps=10),
            )
            second = run_batch_evaluation(
                pipeline,
                scenes,
                root / "batch",
                config=BatchEvaluationConfig(seed=999, steps=60),
            )
            self.assertEqual(pipeline.calls, 2)
            self.assertEqual(second["newly_processed"], 1)
            self.assertEqual(second["reused_successes"], 0)

    def test_batch_resume_invalidates_success_when_artifact_is_missing(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            pipeline = self.FakeBatchPipeline(root / "runs")
            scenes = ["需要检查产物存在性的场景"]
            run_batch_evaluation(pipeline, scenes, root / "batch")
            records = load_batch_results(root / "batch" / "batch_results.jsonl")
            Path(records[0]["raw_image_path"]).unlink()
            second = run_batch_evaluation(pipeline, scenes, root / "batch")
            self.assertEqual(pipeline.calls, 2)
            self.assertEqual(second["newly_processed"], 1)

    def test_batch_continues_after_one_case_fails(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            pipeline = self.FakeBatchPipeline(root / "runs", "失败场景文本")
            summary = run_batch_evaluation(
                pipeline,
                ["正常场景文本", "失败场景文本", "另一个正常场景"],
                root / "batch",
            )

            self.assertEqual(summary["succeeded"], 2)
            self.assertEqual(summary["failed"], 1)
            self.assertFalse(summary["complete"])
            self.assertEqual(pipeline.calls, 3)

    @unittest.skipUnless(
        importlib.util.find_spec("openpyxl"),
        "openpyxl is required for XLSX export verification",
    )
    def test_exporter_fills_template_and_embeds_images(self):
        from openpyxl import Workbook, load_workbook

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            template = root / "template.xlsx"
            template_workbook = Workbook()
            template_worksheet = template_workbook.active
            template_worksheet.title = "prompts"
            template_worksheet.append(
                ["English Prompt", "中文 Prompt", "Image", "Scene"]
            )
            template_workbook.save(template)
            pipeline = self.FakeBatchPipeline(root / "runs")
            batch_dir = root / "batch"
            run_batch_evaluation(
                pipeline,
                ["酒店大堂晨雾天光", "酒店客房黄昏助眠光"],
                batch_dir,
            )
            output = root / "submission.xlsx"
            summary = export_submission(
                batch_dir / "batch_results.jsonl",
                template,
                output,
            )
            workbook = load_workbook(output)
            worksheet = workbook["prompts"]

            self.assertEqual(summary["exported_count"], 2)
            self.assertEqual(worksheet.max_row, 3)
            self.assertEqual(worksheet["D2"].value, "酒店大堂晨雾天光")
            self.assertTrue(worksheet["A3"].value.startswith("English lighting prompt"))
            self.assertEqual(len(worksheet._images), 2)


if __name__ == "__main__":
    unittest.main()
